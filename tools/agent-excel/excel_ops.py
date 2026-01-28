#!/usr/bin/env python3
"""
agent-excel Python operations module
Handles all Excel operations via openpyxl
"""

import sys
import json
import os
import re
from pathlib import Path
from typing import Any, Optional, List, Dict, Tuple

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter, column_index_from_string
    from openpyxl.utils.cell import coordinate_from_string
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.cell.cell import MergedCell
except ImportError:
    print(json.dumps({"success": False, "error": "openpyxl not installed. Run: pip install openpyxl"}))
    sys.exit(1)


class ExcelSession:
    """Manages Excel workbook state"""
    
    def __init__(self, state_file: str):
        self.state_file = state_file
        self.workbook: Optional[Workbook] = None
        self.path: Optional[str] = None
        self.modified: bool = False
        self._load_state()
    
    def _load_state(self):
        """Load session state from file"""
        if os.path.exists(self.state_file):
            try:
                with open(self.state_file, 'r') as f:
                    state = json.load(f)
                if state.get('path') and os.path.exists(state['path']):
                    self.path = state['path']
                    self.workbook = load_workbook(self.path)
                    self.modified = state.get('modified', False)
            except Exception:
                pass
    
    def _save_state(self):
        """Save session state to file"""
        state = {
            'path': self.path,
            'modified': self.modified
        }
        with open(self.state_file, 'w') as f:
            json.dump(state, f)
    
    def _clear_state(self):
        """Clear session state"""
        if os.path.exists(self.state_file):
            os.remove(self.state_file)
        self.workbook = None
        self.path = None
        self.modified = False
    
    @property
    def active_sheet(self):
        if self.workbook:
            return self.workbook.active
        return None
    
    def require_workbook(self):
        """Ensure a workbook is open"""
        if not self.workbook:
            return {"success": False, "error": "No workbook open. Use 'open <path>' first.", "exit_code": 2}
        return None


def success(data: Any) -> dict:
    return {"success": True, "data": data}


def error(msg: str, code: int = 1) -> dict:
    return {"success": False, "error": msg, "exit_code": code}


def parse_range(range_str: str) -> Tuple[str, str]:
    """Parse A1:B2 into (A1, B2) or (A1, A1) for single cell"""
    if ':' in range_str:
        return tuple(range_str.split(':'))
    return (range_str, range_str)


def cell_to_tuple(cell_ref: str) -> Tuple[int, int]:
    """Convert A1 to (row, col) tuple"""
    col_str, row = coordinate_from_string(cell_ref)
    col = column_index_from_string(col_str)
    return (row, col)


def get_cell_value(cell) -> Any:
    """Get cell value, handling merged cells"""
    if isinstance(cell, MergedCell):
        return None
    return cell.value


# =============================================================================
# WORKBOOK COMMANDS
# =============================================================================

def cmd_open(session: ExcelSession, args: List[str]) -> dict:
    """Open an existing workbook"""
    if not args:
        return error("Usage: open <path>", 3)
    
    path = args[0]
    if not os.path.exists(path):
        return error(f"File not found: {path}", 2)
    
    try:
        session.workbook = load_workbook(path)
        session.path = os.path.abspath(path)
        session.modified = False
        session._save_state()
        
        return success({
            "path": session.path,
            "sheets": session.workbook.sheetnames,
            "active_sheet": session.workbook.active.title,
            "modified": False
        })
    except Exception as e:
        return error(f"Failed to open: {e}", 1)


def cmd_new(session: ExcelSession, args: List[str]) -> dict:
    """Create a new workbook"""
    session.workbook = Workbook()
    session.path = args[0] if args else None
    session.modified = True
    session._save_state()
    
    return success({
        "path": session.path,
        "sheets": session.workbook.sheetnames,
        "active_sheet": session.workbook.active.title,
        "modified": True
    })


def cmd_save(session: ExcelSession, args: List[str]) -> dict:
    """Save the workbook"""
    err = session.require_workbook()
    if err:
        return err
    
    path = args[0] if args else session.path
    if not path:
        return error("No path specified. Use 'save <path>'", 3)
    
    try:
        session.workbook.save(path)
        session.path = os.path.abspath(path)
        session.modified = False
        session._save_state()
        return success({"saved": path})
    except PermissionError:
        return error(f"File is locked or permission denied: {path}", 6)
    except Exception as e:
        return error(f"Save failed: {e}", 1)


def cmd_close(session: ExcelSession, args: List[str]) -> dict:
    """Close without saving"""
    session._clear_state()
    return success({"closed": True})


def cmd_status(session: ExcelSession, args: List[str]) -> dict:
    """Get current workbook status"""
    if not session.workbook:
        return success({"open": False})
    
    ws = session.active_sheet
    return success({
        "open": True,
        "path": session.path,
        "sheets": session.workbook.sheetnames,
        "active_sheet": ws.title,
        "modified": session.modified,
        "dimensions": ws.dimensions or "A1"
    })


def cmd_export(session: ExcelSession, args: List[str]) -> dict:
    """Export to CSV or PDF"""
    err = session.require_workbook()
    if err:
        return err
    
    if len(args) < 2:
        return error("Usage: export csv|pdf <path>", 3)
    
    fmt, path = args[0], args[1]
    ws = session.active_sheet
    
    if fmt == "csv":
        import csv
        with open(path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(row)
        return success({"exported": path, "format": "csv", "sheet": ws.title})
    elif fmt == "pdf":
        return error("PDF export requires additional libraries", 1)
    else:
        return error(f"Unknown format: {fmt}. Use csv or pdf", 3)


# =============================================================================
# SHEET COMMANDS
# =============================================================================

def cmd_sheets(session: ExcelSession, args: List[str]) -> dict:
    """List all sheets"""
    err = session.require_workbook()
    if err:
        return err
    
    sheets = []
    for name in session.workbook.sheetnames:
        ws = session.workbook[name]
        sheets.append({
            "name": name,
            "active": ws == session.active_sheet,
            "dimensions": ws.dimensions or "A1"
        })
    return success({"sheets": sheets})


def cmd_sheet(session: ExcelSession, args: List[str]) -> dict:
    """Sheet operations: switch, new, rename, delete, copy"""
    err = session.require_workbook()
    if err:
        return err
    
    if not args:
        return success({"active_sheet": session.active_sheet.title})
    
    subcmd = args[0]
    
    if subcmd == "new":
        if len(args) < 2:
            return error("Usage: sheet new <name>", 3)
        name = args[1]
        session.workbook.create_sheet(name)
        session.workbook.active = session.workbook[name]
        session.modified = True
        session._save_state()
        return success({"created": name, "active_sheet": name})
    
    elif subcmd == "rename":
        if len(args) < 3:
            return error("Usage: sheet rename <old> <new>", 3)
        old, new = args[1], args[2]
        if old not in session.workbook.sheetnames:
            return error(f"Sheet not found: {old}", 4)
        session.workbook[old].title = new
        session.modified = True
        session._save_state()
        return success({"renamed": old, "to": new})
    
    elif subcmd == "delete":
        if len(args) < 2:
            return error("Usage: sheet delete <name>", 3)
        name = args[1]
        if name not in session.workbook.sheetnames:
            return error(f"Sheet not found: {name}", 4)
        del session.workbook[name]
        session.modified = True
        session._save_state()
        return success({"deleted": name})
    
    elif subcmd == "copy":
        if len(args) < 2:
            return error("Usage: sheet copy <source> [dest_name]", 3)
        src = args[1]
        if src not in session.workbook.sheetnames:
            return error(f"Sheet not found: {src}", 4)
        dest = args[2] if len(args) > 2 else f"{src} Copy"
        session.workbook.copy_worksheet(session.workbook[src]).title = dest
        session.modified = True
        session._save_state()
        return success({"copied": src, "to": dest})
    
    else:
        # Switch to sheet
        if subcmd not in session.workbook.sheetnames:
            return error(f"Sheet not found: {subcmd}", 4)
        session.workbook.active = session.workbook[subcmd]
        session._save_state()
        return success({"active_sheet": subcmd})


# =============================================================================
# READING COMMANDS
# =============================================================================

def cmd_snapshot(session: ExcelSession, args: List[str]) -> dict:
    """Get overview of active sheet"""
    err = session.require_workbook()
    if err:
        return err
    
    ws = session.active_sheet
    
    # Parse options
    range_str = None
    used_only = False
    with_headers = False
    
    i = 0
    while i < len(args):
        if args[i] == "--range" and i + 1 < len(args):
            range_str = args[i + 1]
            i += 2
        elif args[i] == "--used":
            used_only = True
            i += 1
        elif args[i] == "--headers":
            with_headers = True
            i += 1
        else:
            i += 1
    
    # Get dimensions
    dims = ws.dimensions or "A1:A1"
    min_row, max_row = ws.min_row or 1, ws.max_row or 1
    min_col, max_col = ws.min_column or 1, ws.max_column or 1
    
    result = {
        "sheet": ws.title,
        "used_range": dims,
        "row_count": max_row - min_row + 1,
        "col_count": max_col - min_col + 1
    }
    
    # Get headers (first row)
    if with_headers or True:  # Always include headers
        headers = []
        for col in range(min_col, max_col + 1):
            val = ws.cell(row=min_row, column=col).value
            headers.append(str(val) if val is not None else f"Col{get_column_letter(col)}")
        result["headers"] = headers
    
    # Get preview rows
    preview = []
    preview_count = 5
    for row_idx in range(min_row + 1, min(min_row + 1 + preview_count, max_row + 1)):
        row_data = {"row": row_idx}
        for col in range(min_col, max_col + 1):
            col_letter = get_column_letter(col)
            val = ws.cell(row=row_idx, column=col).value
            row_data[col_letter] = val
        preview.append(row_data)
    
    result["preview"] = preview
    result["preview_rows"] = len(preview)
    
    return success(result)


def cmd_get(session: ExcelSession, args: List[str]) -> dict:
    """Get cell or range values"""
    err = session.require_workbook()
    if err:
        return err
    
    if not args:
        return error("Usage: get <cell|range|row N|col L|formula cell>", 3)
    
    ws = session.active_sheet
    
    # Handle special forms
    if args[0] == "row":
        if len(args) < 2:
            return error("Usage: get row <number>", 3)
        row_num = int(args[1])
        values = []
        for col in range(1, ws.max_column + 1):
            values.append(ws.cell(row=row_num, column=col).value)
        return success({"row": row_num, "values": values})
    
    elif args[0] == "col":
        if len(args) < 2:
            return error("Usage: get col <letter>", 3)
        col_letter = args[1].upper()
        col_num = column_index_from_string(col_letter)
        values = []
        for row in range(1, ws.max_row + 1):
            values.append(ws.cell(row=row, column=col_num).value)
        return success({"col": col_letter, "values": values})
    
    elif args[0] == "formula":
        if len(args) < 2:
            return error("Usage: get formula <cell>", 3)
        cell_ref = args[1].upper()
        cell = ws[cell_ref]
        return success({
            "cell": cell_ref,
            "formula": cell.value if isinstance(cell.value, str) and cell.value.startswith('=') else None,
            "value": cell.value
        })
    
    # Regular cell or range
    cell_ref = args[0].upper()
    
    if ':' in cell_ref:
        # Range
        start, end = parse_range(cell_ref)
        start_row, start_col = cell_to_tuple(start)
        end_row, end_col = cell_to_tuple(end)
        
        values = []
        formulas = []
        for row in range(start_row, end_row + 1):
            row_vals = []
            row_forms = []
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                row_vals.append(get_cell_value(cell))
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    row_forms.append(cell.value)
                else:
                    row_forms.append(None)
            values.append(row_vals)
            formulas.append(row_forms)
        
        return success({"range": cell_ref, "values": values, "formulas": formulas})
    else:
        # Single cell
        cell = ws[cell_ref]
        return success({
            "cell": cell_ref,
            "value": get_cell_value(cell),
            "formula": cell.value if isinstance(cell.value, str) and cell.value.startswith('=') else None
        })


# =============================================================================
# WRITING COMMANDS
# =============================================================================

def cmd_set(session: ExcelSession, args: List[str]) -> dict:
    """Set cell or range values"""
    err = session.require_workbook()
    if err:
        return err
    
    if len(args) < 2:
        return error("Usage: set <cell|range> <value|json_array>", 3)
    
    ws = session.active_sheet
    cell_ref = args[0].upper()
    value = ' '.join(args[1:])
    
    # Try to parse as JSON for range setting
    try:
        parsed = json.loads(value)
        if isinstance(parsed, list):
            # 2D array for range
            start, _ = parse_range(cell_ref) if ':' in cell_ref else (cell_ref, cell_ref)
            start_row, start_col = cell_to_tuple(start)
            
            for r_idx, row_data in enumerate(parsed):
                if isinstance(row_data, list):
                    for c_idx, val in enumerate(row_data):
                        ws.cell(row=start_row + r_idx, column=start_col + c_idx, value=val)
                else:
                    ws.cell(row=start_row + r_idx, column=start_col, value=row_data)
            
            session.modified = True
            session._save_state()
            return success({"set": cell_ref, "rows": len(parsed)})
    except json.JSONDecodeError:
        pass
    
    # Single value - try type coercion
    if value.lower() == 'true':
        value = True
    elif value.lower() == 'false':
        value = False
    elif value.lower() in ('null', 'none', ''):
        value = None
    else:
        try:
            if '.' in value:
                value = float(value)
            else:
                value = int(value)
        except ValueError:
            pass  # Keep as string
    
    ws[cell_ref] = value
    session.modified = True
    session._save_state()
    return success({"set": cell_ref, "value": value})


def cmd_fill(session: ExcelSession, args: List[str]) -> dict:
    """Fill range with value"""
    err = session.require_workbook()
    if err:
        return err
    
    if len(args) < 2:
        return error("Usage: fill <range> <value>", 3)
    
    ws = session.active_sheet
    range_str = args[0].upper()
    value = ' '.join(args[1:])
    
    # Type coercion
    try:
        if '.' in value:
            value = float(value)
        else:
            value = int(value)
    except ValueError:
        pass
    
    start, end = parse_range(range_str)
    start_row, start_col = cell_to_tuple(start)
    end_row, end_col = cell_to_tuple(end)
    
    count = 0
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            ws.cell(row=row, column=col, value=value)
            count += 1
    
    session.modified = True
    session._save_state()
    return success({"filled": range_str, "value": value, "cells": count})


def cmd_clear(session: ExcelSession, args: List[str]) -> dict:
    """Clear cell contents"""
    err = session.require_workbook()
    if err:
        return err
    
    if not args:
        return error("Usage: clear <range> [--all]", 3)
    
    ws = session.active_sheet
    range_str = args[0].upper()
    clear_all = "--all" in args
    
    start, end = parse_range(range_str)
    start_row, start_col = cell_to_tuple(start)
    end_row, end_col = cell_to_tuple(end)
    
    count = 0
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.value = None
            if clear_all:
                cell.font = Font()
                cell.fill = PatternFill()
                cell.alignment = Alignment()
            count += 1
    
    session.modified = True
    session._save_state()
    return success({"cleared": range_str, "cells": count, "format_cleared": clear_all})


def cmd_formula(session: ExcelSession, args: List[str]) -> dict:
    """Set formula in cell"""
    err = session.require_workbook()
    if err:
        return err
    
    if len(args) < 2:
        return error("Usage: formula <cell> <expression>", 3)
    
    ws = session.active_sheet
    cell_ref = args[0].upper()
    formula = ' '.join(args[1:])
    
    if not formula.startswith('='):
        formula = '=' + formula
    
    ws[cell_ref] = formula
    session.modified = True
    session._save_state()
    return success({"cell": cell_ref, "formula": formula})


# =============================================================================
# STRUCTURE COMMANDS
# =============================================================================

def cmd_insert(session: ExcelSession, args: List[str]) -> dict:
    """Insert rows or columns"""
    err = session.require_workbook()
    if err:
        return err
    
    if len(args) < 2:
        return error("Usage: insert row|col <position> [count]", 3)
    
    ws = session.active_sheet
    what = args[0]
    pos = args[1]
    count = int(args[2]) if len(args) > 2 else 1
    
    if what == "row":
        row_num = int(pos)
        ws.insert_rows(row_num, count)
        session.modified = True
        session._save_state()
        return success({"inserted": "rows", "at": row_num, "count": count})
    
    elif what == "col":
        col_num = column_index_from_string(pos.upper())
        ws.insert_cols(col_num, count)
        session.modified = True
        session._save_state()
        return success({"inserted": "cols", "at": pos.upper(), "count": count})
    
    return error(f"Unknown: {what}. Use row or col", 3)


def cmd_delete(session: ExcelSession, args: List[str]) -> dict:
    """Delete rows or columns"""
    err = session.require_workbook()
    if err:
        return err
    
    if len(args) < 2:
        return error("Usage: delete row|col <position> [count]", 3)
    
    ws = session.active_sheet
    what = args[0]
    pos = args[1]
    count = int(args[2]) if len(args) > 2 else 1
    
    if what == "row":
        row_num = int(pos)
        ws.delete_rows(row_num, count)
        session.modified = True
        session._save_state()
        return success({"deleted": "rows", "at": row_num, "count": count})
    
    elif what == "col":
        col_num = column_index_from_string(pos.upper())
        ws.delete_cols(col_num, count)
        session.modified = True
        session._save_state()
        return success({"deleted": "cols", "at": pos.upper(), "count": count})
    
    return error(f"Unknown: {what}. Use row or col", 3)


def cmd_autofit(session: ExcelSession, args: List[str]) -> dict:
    """Auto-fit column width"""
    err = session.require_workbook()
    if err:
        return err
    
    if len(args) < 2:
        return error("Usage: autofit col <letter>", 3)
    
    ws = session.active_sheet
    
    if args[0] == "col":
        col_letter = args[1].upper()
        col_num = column_index_from_string(col_letter)
        
        max_length = 0
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_num)
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        
        ws.column_dimensions[col_letter].width = max_length + 2
        session.modified = True
        session._save_state()
        return success({"autofit": col_letter, "width": max_length + 2})
    
    return error("Usage: autofit col <letter>", 3)


# =============================================================================
# FIND/FILTER COMMANDS
# =============================================================================

def cmd_find(session: ExcelSession, args: List[str]) -> dict:
    """Find text in sheet"""
    err = session.require_workbook()
    if err:
        return err
    
    if not args:
        return error("Usage: find <text> [--col <letter>]", 3)
    
    ws = session.active_sheet
    query = args[0]
    col_filter = None
    
    # Parse --col option
    if "--col" in args:
        idx = args.index("--col")
        if idx + 1 < len(args):
            col_filter = column_index_from_string(args[idx + 1].upper())
    
    matches = []
    for row in range(1, ws.max_row + 1):
        cols_to_search = [col_filter] if col_filter else range(1, ws.max_column + 1)
        for col in cols_to_search:
            cell = ws.cell(row=row, column=col)
            if cell.value and query.lower() in str(cell.value).lower():
                matches.append({
                    "cell": f"{get_column_letter(col)}{row}",
                    "value": cell.value
                })
    
    return success({"query": query, "matches": matches, "count": len(matches)})


def cmd_replace(session: ExcelSession, args: List[str]) -> dict:
    """Replace text in sheet"""
    err = session.require_workbook()
    if err:
        return err
    
    if len(args) < 2:
        return error("Usage: replace <old> <new>", 3)
    
    ws = session.active_sheet
    old_text, new_text = args[0], args[1]
    
    count = 0
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value and isinstance(cell.value, str) and old_text in cell.value:
                cell.value = cell.value.replace(old_text, new_text)
                count += 1
    
    if count > 0:
        session.modified = True
        session._save_state()
    
    return success({"replaced": old_text, "with": new_text, "count": count})


def cmd_filter(session: ExcelSession, args: List[str]) -> dict:
    """Filter rows by column value"""
    err = session.require_workbook()
    if err:
        return err
    
    if not args:
        return error("Usage: filter <col> <op> <value> | filter clear", 3)
    
    ws = session.active_sheet
    
    if args[0] == "clear":
        ws.auto_filter.ref = None
        session.modified = True
        session._save_state()
        return success({"filter": "cleared"})
    
    if len(args) < 3:
        return error("Usage: filter <col> <op> <value>", 3)
    
    col_letter = args[0].upper()
    op = args[1]
    value = ' '.join(args[2:])
    
    # Type coercion for comparison
    try:
        value = float(value)
    except ValueError:
        pass
    
    col_num = column_index_from_string(col_letter)
    matches = []
    
    for row in range(2, ws.max_row + 1):  # Skip header
        cell_val = ws.cell(row=row, column=col_num).value
        
        match = False
        if op == "=" and cell_val == value:
            match = True
        elif op == "!=" and cell_val != value:
            match = True
        elif op == ">" and cell_val is not None and cell_val > value:
            match = True
        elif op == "<" and cell_val is not None and cell_val < value:
            match = True
        elif op == ">=" and cell_val is not None and cell_val >= value:
            match = True
        elif op == "<=" and cell_val is not None and cell_val <= value:
            match = True
        elif op == "contains" and cell_val and str(value).lower() in str(cell_val).lower():
            match = True
        elif op == "empty" and (cell_val is None or cell_val == ""):
            match = True
        elif op == "notempty" and cell_val is not None and cell_val != "":
            match = True
        
        if match:
            row_data = {"row": row}
            for c in range(1, ws.max_column + 1):
                row_data[get_column_letter(c)] = ws.cell(row=row, column=c).value
            matches.append(row_data)
    
    return success({
        "filter": {"col": col_letter, "op": op, "value": value},
        "matches": matches,
        "count": len(matches)
    })


def cmd_sort(session: ExcelSession, args: List[str]) -> dict:
    """Sort by column"""
    err = session.require_workbook()
    if err:
        return err
    
    if not args:
        return error("Usage: sort <col> [asc|desc]", 3)
    
    ws = session.active_sheet
    col_letter = args[0].upper()
    order = args[1].lower() if len(args) > 1 else "asc"
    reverse = order == "desc"
    
    col_num = column_index_from_string(col_letter)
    
    # Get all data rows (skip header)
    data = []
    for row in range(2, ws.max_row + 1):
        row_data = []
        for col in range(1, ws.max_column + 1):
            row_data.append(ws.cell(row=row, column=col).value)
        data.append(row_data)
    
    # Sort by the specified column
    def sort_key(row):
        val = row[col_num - 1]
        if val is None:
            return (1, "")  # None values last
        return (0, val)
    
    data.sort(key=sort_key, reverse=reverse)
    
    # Write back
    for r_idx, row_data in enumerate(data):
        for c_idx, val in enumerate(row_data):
            ws.cell(row=r_idx + 2, column=c_idx + 1, value=val)
    
    session.modified = True
    session._save_state()
    return success({"sorted_by": col_letter, "order": order, "rows": len(data)})


# =============================================================================
# FORMAT COMMANDS
# =============================================================================

def cmd_format(session: ExcelSession, args: List[str]) -> dict:
    """Format cells"""
    err = session.require_workbook()
    if err:
        return err
    
    if not args:
        return error("Usage: format <range> --bold --italic --color HEX --bg HEX --align l|c|r", 3)
    
    ws = session.active_sheet
    range_str = args[0].upper()
    
    # Parse options
    bold = "--bold" in args
    italic = "--italic" in args
    underline = "--underline" in args
    color = None
    bg = None
    align = None
    number_format = None
    
    i = 1
    while i < len(args):
        if args[i] == "--color" and i + 1 < len(args):
            color = args[i + 1].lstrip('#')
            i += 2
        elif args[i] == "--bg" and i + 1 < len(args):
            bg = args[i + 1].lstrip('#')
            i += 2
        elif args[i] == "--align" and i + 1 < len(args):
            align = {"l": "left", "c": "center", "r": "right"}.get(args[i + 1], args[i + 1])
            i += 2
        elif args[i] == "--number" and i + 1 < len(args):
            number_format = args[i + 1]
            i += 2
        else:
            i += 1
    
    start, end = parse_range(range_str)
    start_row, start_col = cell_to_tuple(start)
    end_row, end_col = cell_to_tuple(end)
    
    count = 0
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            
            # Build font
            font_kwargs = {}
            if bold:
                font_kwargs['bold'] = True
            if italic:
                font_kwargs['italic'] = True
            if underline:
                font_kwargs['underline'] = 'single'
            if color:
                font_kwargs['color'] = color
            if font_kwargs:
                cell.font = Font(**font_kwargs)
            
            # Background
            if bg:
                cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
            
            # Alignment
            if align:
                cell.alignment = Alignment(horizontal=align)
            
            # Number format
            if number_format:
                cell.number_format = number_format
            
            count += 1
    
    session.modified = True
    session._save_state()
    return success({"formatted": range_str, "cells": count})


def cmd_merge(session: ExcelSession, args: List[str]) -> dict:
    """Merge cells"""
    err = session.require_workbook()
    if err:
        return err
    
    if not args:
        return error("Usage: merge <range>", 3)
    
    ws = session.active_sheet
    range_str = args[0].upper()
    
    ws.merge_cells(range_str)
    session.modified = True
    session._save_state()
    return success({"merged": range_str})


def cmd_unmerge(session: ExcelSession, args: List[str]) -> dict:
    """Unmerge cells"""
    err = session.require_workbook()
    if err:
        return err
    
    if not args:
        return error("Usage: unmerge <range>", 3)
    
    ws = session.active_sheet
    range_str = args[0].upper()
    
    ws.unmerge_cells(range_str)
    session.modified = True
    session._save_state()
    return success({"unmerged": range_str})


# =============================================================================
# FORMULA HELPERS
# =============================================================================

def cmd_sum(session: ExcelSession, args: List[str]) -> dict:
    """Calculate SUM of range"""
    err = session.require_workbook()
    if err:
        return err
    
    if not args:
        return error("Usage: sum <range>", 3)
    
    ws = session.active_sheet
    range_str = args[0].upper()
    
    total = 0
    start, end = parse_range(range_str)
    start_row, start_col = cell_to_tuple(start)
    end_row, end_col = cell_to_tuple(end)
    
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            val = ws.cell(row=row, column=col).value
            if isinstance(val, (int, float)):
                total += val
    
    return success({"range": range_str, "sum": total})


def cmd_avg(session: ExcelSession, args: List[str]) -> dict:
    """Calculate AVERAGE of range"""
    err = session.require_workbook()
    if err:
        return err
    
    if not args:
        return error("Usage: avg <range>", 3)
    
    ws = session.active_sheet
    range_str = args[0].upper()
    
    total = 0
    count = 0
    start, end = parse_range(range_str)
    start_row, start_col = cell_to_tuple(start)
    end_row, end_col = cell_to_tuple(end)
    
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            val = ws.cell(row=row, column=col).value
            if isinstance(val, (int, float)):
                total += val
                count += 1
    
    avg = total / count if count > 0 else 0
    return success({"range": range_str, "average": avg, "count": count})


def cmd_count(session: ExcelSession, args: List[str]) -> dict:
    """Count non-empty cells"""
    err = session.require_workbook()
    if err:
        return err
    
    if not args:
        return error("Usage: count <range>", 3)
    
    ws = session.active_sheet
    range_str = args[0].upper()
    
    count = 0
    start, end = parse_range(range_str)
    start_row, start_col = cell_to_tuple(start)
    end_row, end_col = cell_to_tuple(end)
    
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                count += 1
    
    return success({"range": range_str, "count": count})


def cmd_max(session: ExcelSession, args: List[str]) -> dict:
    """Find MAX in range"""
    err = session.require_workbook()
    if err:
        return err
    
    if not args:
        return error("Usage: max <range>", 3)
    
    ws = session.active_sheet
    range_str = args[0].upper()
    
    values = []
    start, end = parse_range(range_str)
    start_row, start_col = cell_to_tuple(start)
    end_row, end_col = cell_to_tuple(end)
    
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            val = ws.cell(row=row, column=col).value
            if isinstance(val, (int, float)):
                values.append(val)
    
    max_val = max(values) if values else None
    return success({"range": range_str, "max": max_val})


def cmd_min(session: ExcelSession, args: List[str]) -> dict:
    """Find MIN in range"""
    err = session.require_workbook()
    if err:
        return err
    
    if not args:
        return error("Usage: min <range>", 3)
    
    ws = session.active_sheet
    range_str = args[0].upper()
    
    values = []
    start, end = parse_range(range_str)
    start_row, start_col = cell_to_tuple(start)
    end_row, end_col = cell_to_tuple(end)
    
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            val = ws.cell(row=row, column=col).value
            if isinstance(val, (int, float)):
                values.append(val)
    
    min_val = min(values) if values else None
    return success({"range": range_str, "min": min_val})


# =============================================================================
# MAIN DISPATCH
# =============================================================================

COMMANDS = {
    'open': cmd_open,
    'new': cmd_new,
    'save': cmd_save,
    'close': cmd_close,
    'status': cmd_status,
    'export': cmd_export,
    'sheets': cmd_sheets,
    'sheet': cmd_sheet,
    'snapshot': cmd_snapshot,
    'get': cmd_get,
    'set': cmd_set,
    'fill': cmd_fill,
    'clear': cmd_clear,
    'formula': cmd_formula,
    'insert': cmd_insert,
    'delete': cmd_delete,
    'autofit': cmd_autofit,
    'find': cmd_find,
    'replace': cmd_replace,
    'filter': cmd_filter,
    'sort': cmd_sort,
    'format': cmd_format,
    'merge': cmd_merge,
    'unmerge': cmd_unmerge,
    'sum': cmd_sum,
    'avg': cmd_avg,
    'count': cmd_count,
    'max': cmd_max,
    'min': cmd_min,
}


def main():
    if len(sys.argv) < 3:
        print(json.dumps(error("No command provided")))
        sys.exit(1)
    
    state_file = sys.argv[1]
    cmd = sys.argv[2]
    args = sys.argv[3:]
    
    if cmd not in COMMANDS:
        # Fuzzy suggestion
        suggestions = [c for c in COMMANDS if c.startswith(cmd)]
        if suggestions:
            result = error(f"Unknown command: {cmd}. Did you mean: {', '.join(suggestions)}?")
        else:
            result = error(f"Unknown command: {cmd}")
        print(json.dumps(result))
        sys.exit(1)
    
    session = ExcelSession(state_file)
    result = COMMANDS[cmd](session, args)
    
    print(json.dumps(result, default=str))
    sys.exit(result.get('exit_code', 0) if not result.get('success', True) else 0)


if __name__ == '__main__':
    main()
